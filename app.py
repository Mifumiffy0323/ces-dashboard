#!/usr/bin/env python3
"""
CES Dashboard — Flask version for Render deployment
Local:  python app.py
Deploy: Render.com (see README)
"""

import json
import os
import secrets
import threading

import pandas as pd
from flask import (Flask, Response, jsonify, redirect, render_template_string,
                   request, session, url_for)
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

_excel_lock = threading.Lock()

# ── 許可メールアドレス ──────────────────────────────────────────────────────────
_raw = os.environ.get("ALLOWED_EMAILS", "")
ALLOWED_EMAILS = {e.strip().lower() for e in _raw.split(",") if e.strip()}

# ── データ読み込み ──────────────────────────────────────────────────────────────

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "Mater_PythonDataFromSnowflake.xlsx")

FILTER_COLS = [
    "Product", "ANC", "Network", "Case_Size", "Band_Size",
    "Screen_Type", "Screen_Size", "Chip", "CPU", "GPU",
    "Unified_Memory", "SSD", "StockAvailability", "Lists_Colour",
    "Robot_Status", "Band_Colour", "Partner",
]
VALUE_COLS = [
    "PriceMinusPoints", "Selling_Price", "Lists_Price",
    "Pointback", "Hidden_Discount", "Displayed_Discount", "Total_Discount",
]
DISPLAY_COLS = [
    "Date", "Partner", "Product", "StockAvailability",
    "Robot_Status", "Selling_Price", "Lists_Price", "Pointback",
    "Hidden_Discount", "Displayed_Discount", "Total_Discount", "PriceMinusPoints",
]


def load_data():
    print("Loading Excel data …")
    cols_needed = [
        "TASK_CREATED_AT", "CREATED_TS", "ROBOT_NAME", "MONITOR_NAME",
        "Product", "ANC", "Network", "Case_Size", "Band_Size",
        "Screen_Type", "Screen_Size", "Chip", "CPU", "GPU",
        "Unified_Memory", "SSD", "StockAvailability", "Lists_Colour",
        "Robot_Status", "Band_Colour",
        "PriceMinusPoints", "Selling_Price", "Lists_Price",
        "Pointback", "Hidden_Discount", "Displayed_Discount", "Total_Discount",
    ]
    df = pd.read_excel(EXCEL_FILE, sheet_name="CES",
                       usecols=lambda c: c in cols_needed)
    partner_df = pd.read_excel(EXCEL_FILE, sheet_name="CESPartner")
    partner_map = dict(zip(partner_df["ROBOT_NAME"], partner_df["Partner"]))
    df["Partner"] = df["ROBOT_NAME"].map(partner_map).fillna("-")
    df["Date"] = pd.to_datetime(df["CREATED_TS"], format="mixed").dt.strftime("%Y-%m-%d")

    for col in VALUE_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    for col in FILTER_COLS + ["MONITOR_NAME"]:
        df[col] = df[col].fillna("-").astype(str)

    print(f"Loaded {len(df):,} rows. Date range: {df['Date'].min()} → {df['Date'].max()}")
    return df


def load_comments():
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="Comments")
        return df.to_dict(orient="records")
    except Exception:
        return []


def save_comment(partner, date, price, text):
    global COMMENTS
    updated = False
    for c in COMMENTS:
        if c.get("Partner") == partner and c.get("Date") == date:
            c["Comment"] = text; c["Price"] = price; updated = True; break
    if not updated:
        COMMENTS.append({"Partner": partner, "Date": date, "Price": price, "Comment": text})
    def _write():
        with _excel_lock:
            try:
                wb = load_workbook(EXCEL_FILE)
                if "Comments" in wb.sheetnames: del wb["Comments"]
                ws = wb.create_sheet("Comments")
                ws.append(["Partner", "Date", "Price", "Comment"])
                for c in COMMENTS:
                    ws.append([c.get("Partner",""), c.get("Date",""), c.get("Price",0), c.get("Comment","")])
                wb.save(EXCEL_FILE)
            except Exception as e:
                print(f"Comment save error: {e}")
    threading.Thread(target=_write, daemon=True).start()


def delete_comment(partner, date):
    global COMMENTS
    COMMENTS = [c for c in COMMENTS
                if not (c.get("Partner") == partner and c.get("Date") == date)]
    def _write():
        with _excel_lock:
            try:
                wb = load_workbook(EXCEL_FILE)
                if "Comments" in wb.sheetnames: del wb["Comments"]
                if COMMENTS:
                    ws = wb.create_sheet("Comments")
                    ws.append(["Partner", "Date", "Price", "Comment"])
                    for c in COMMENTS:
                        ws.append([c.get("Partner",""), c.get("Date",""), c.get("Price",0), c.get("Comment","")])
                wb.save(EXCEL_FILE)
            except Exception as e:
                print(f"Comment delete error: {e}")
    threading.Thread(target=_write, daemon=True).start()


def load_bd_focus():
    try:
        bf = pd.read_excel(os.path.join(BASE_DIR, "20260216_BDFocus.xlsx"))
        bf = bf[bf["BDFocus"] == "Yes"]
        conditions = []
        for _, row in bf.iterrows():
            cond = {"Product": str(row["Product"])}
            if pd.notna(row.get("Chip")) and str(row["Chip"]) not in ("", "nan"):
                cond["Chip"] = str(row["Chip"])
            if pd.notna(row.get("SSD")) and str(row["SSD"]) not in ("", "nan"):
                cond["SSD"] = str(row["SSD"])
            if pd.notna(row.get("ScreenSize")) and str(row["ScreenSize"]) not in ("", "nan"):
                cond["Screen_Size"] = str(row["ScreenSize"])
            conditions.append(cond)
        print(f"BD Focus: {len(conditions)} products loaded")
        return conditions
    except Exception as e:
        print(f"BD Focus load error: {e}")
        return []


DF        = load_data()
ALL_DATES = sorted(DF["Date"].unique())
DATE_MIN  = ALL_DATES[0]
DATE_MAX  = ALL_DATES[-1]
DATE_15   = ALL_DATES[-15] if len(ALL_DATES) >= 15 else DATE_MIN

OPTIONS = {col: sorted(DF[col].unique().tolist()) for col in FILTER_COLS}
OPTIONS["Date"] = ALL_DATES
COMMENTS = load_comments()
BD_FOCUS_CONDITIONS = load_bd_focus()


# ── 認証ヘルパー ────────────────────────────────────────────────────────────────

def is_logged_in():
    return session.get("email") in ALLOWED_EMAILS


def require_login(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not is_logged_in():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


# ── フィルター適用 ──────────────────────────────────────────────────────────────

def apply_filters(params):
    df = DF.copy()
    date_from = params.get("date_from", DATE_15)
    date_to   = params.get("date_to",   DATE_MAX)
    if isinstance(date_from, list): date_from = date_from[0]
    if isinstance(date_to,   list): date_to   = date_to[0]
    df = df[(df["Date"] >= date_from) & (df["Date"] <= date_to)]
    for col in FILTER_COLS:
        vals = params.getlist(col) if hasattr(params, "getlist") else params.get(col, [])
        if vals and vals != [""]:
            df = df[df[col].isin(vals)]
    if params.get("bd_focus") == "true" and BD_FOCUS_CONDITIONS:
        mask = pd.Series(False, index=df.index)
        for cond in BD_FOCUS_CONDITIONS:
            m = pd.Series(True, index=df.index)
            for col, val in cond.items():
                m = m & (df[col] == val)
            mask = mask | m
        df = df[mask]
    return df


# ── ログイン画面 ────────────────────────────────────────────────────────────────

LOGIN_HTML = """<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CES Dashboard — Login</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #f4f6f9; display: flex; align-items: center;
    justify-content: center; min-height: 100vh;
  }
  .card {
    background: #fff; border-radius: 16px; padding: 44px 48px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.10); width: 100%; max-width: 400px;
    text-align: center;
  }
  .logo { font-size: 26px; font-weight: 800; color: #141829; margin-bottom: 6px; }
  .logo span { color: #4f8ef7; }
  .subtitle { font-size: 13px; color: #8a93a8; margin-bottom: 32px; }
  label { display: block; text-align: left; font-size: 12px;
          font-weight: 600; color: #4a5568; margin-bottom: 6px; }
  input[type=email] {
    width: 100%; padding: 11px 14px; border: 1px solid #e0e5ef;
    border-radius: 8px; font-size: 14px; color: #2c3e50;
    outline: none; transition: border-color 0.15s;
  }
  input[type=email]:focus { border-color: #4f8ef7; }
  button {
    width: 100%; margin-top: 20px; padding: 12px;
    background: #4f8ef7; color: #fff; border: none;
    border-radius: 8px; font-size: 15px; font-weight: 600;
    cursor: pointer; transition: background 0.15s;
  }
  button:hover { background: #3a7de0; }
  .error {
    margin-top: 16px; padding: 10px 14px; background: #fde8e8;
    border-radius: 8px; color: #c0392b; font-size: 13px;
  }
  .note { margin-top: 20px; font-size: 11px; color: #aaa; }
</style>
</head>
<body>
<div class="card">
  <div class="logo">CES <span>Dashboard</span></div>
  <div class="subtitle">社内限定 — メールアドレスでアクセス</div>
  <form method="POST" action="/login">
    <label for="email">メールアドレス</label>
    <input type="email" id="email" name="email"
           placeholder="you@company.com" required autofocus>
    {% if error %}
    <div class="error">{{ error }}</div>
    {% endif %}
    <button type="submit">アクセスする</button>
  </form>
  <div class="note">登録済みメールアドレスのみアクセスできます</div>
</div>
</body>
</html>
"""


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if email in ALLOWED_EMAILS:
            session["email"] = email
            return redirect(url_for("index"))
        error = "このメールアドレスはアクセス権がありません。"
    return render_template_string(LOGIN_HTML, error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── メインページ ────────────────────────────────────────────────────────────────

@app.route("/")
@require_login
def index():
    html = (HTML
            .replace("DATE_MAX_PLACEHOLDER", DATE_MAX)
            .replace("DATE_15_PLACEHOLDER",  DATE_15)
            .replace("DATE_MAX_JS_PLACEHOLDER", DATE_MAX)
            .replace("DATE_MIN_JS_PLACEHOLDER", DATE_MIN)
            .replace("DATE_15_JS_PLACEHOLDER",  DATE_15)
            .replace("USER_EMAIL_PLACEHOLDER", session.get("email", "")))
    return Response(html, mimetype="text/html; charset=utf-8")


# ── API ─────────────────────────────────────────────────────────────────────────

@app.route("/api/options")
@require_login
def api_options():
    return jsonify(OPTIONS)


@app.route("/api/data")
@require_login
def api_data():
    params  = request.args
    param   = params.get("param", "Selling_Price")
    if param not in VALUE_COLS:
        param = "Selling_Price"

    df = apply_filters(params)

    if df.empty:
        return jsonify({
            "kpis": {"count": 0, "avg": 0, "min": 0, "max": 0, "param": param},
            "by_partner": [], "by_product": [], "by_date": [], "table": [],
            "total_rows": 0, "table_mode": "latest",
            "available_options": {},
        })

    valid = df[df[param] > 0][param]
    kpis = {
        "count": int(len(df)),
        "avg":   int(valid.mean()) if len(valid) else 0,
        "min":   int(valid.min())  if len(valid) else 0,
        "max":   int(valid.max())  if len(valid) else 0,
        "param": param,
    }

    by_partner = (
        df[df[param] > 0].groupby("Partner")[param]
        .agg(avg="mean", count="count").reset_index()
        .sort_values("avg", ascending=False)
    )
    by_partner["avg"] = by_partner["avg"].round(0).astype(int)

    by_product = (
        df[df[param] > 0].groupby("Product")[param]
        .agg(avg="mean", count="count").reset_index()
        .sort_values("avg", ascending=False).head(25)
    )
    by_product["avg"] = by_product["avg"].round(0).astype(int)

    by_date = (
        df[df[param] > 0].groupby("Date")[param].mean()
        .reset_index().rename(columns={param: "avg"})
    )
    by_date["avg"] = by_date["avg"].round(0).astype(int)

    partner_filter = [v for v in params.getlist("Partner") if v]
    latest_date = df["Date"].max()
    if partner_filter:
        table_df = (df.sort_values("Date", ascending=False)
                      .drop_duplicates(subset=["Partner", "Product"])[DISPLAY_COLS]
                      .sort_values(["Partner", "Product"]))
        table_mode = "per_product"
    else:
        table_df = df[df["Date"] == latest_date][DISPLAY_COLS].sort_values(["Partner", "Product"])
        table_mode = "latest"

    avail = {col: sorted(df[col].unique().tolist()) for col in FILTER_COLS}

    return jsonify({
        "kpis":             kpis,
        "by_partner":       by_partner.to_dict(orient="records"),
        "by_product":       by_product.to_dict(orient="records"),
        "by_date":          by_date.to_dict(orient="records"),
        "table":            table_df.to_dict(orient="records"),
        "total_rows":       int(len(df)),
        "table_mode":       table_mode,
        "available_options": avail,
    })


@app.route("/api/pivot")
@require_login
def api_pivot():
    params = request.args
    param  = params.get("param", "Selling_Price")
    if param not in VALUE_COLS:
        param = "Selling_Price"

    df = apply_filters(params)
    if df.empty:
        return jsonify({"dates": [], "partners": [], "values": {},
                        "stock": {}, "changes": {}, "param": param, "best_prices": {}})

    all_dates = sorted(df["Date"].unique().tolist())
    dates     = all_dates[-15:]
    partners  = sorted(p for p in df["Partner"].unique().tolist() if p != "-")
    prev_date = all_dates[-16] if len(all_dates) > 15 else None
    calc_dates = ([prev_date] if prev_date else []) + dates
    df_calc = df[df["Date"].isin(calc_dates)]

    pivot_vals, stock_vals = {}, {}
    for (partner, date), group in df_calc.groupby(["Partner", "Date"]):
        if partner == "-":
            continue
        if partner not in pivot_vals:
            pivot_vals[partner] = {}
            stock_vals[partner] = {}
        valid = group[group[param] > 0][param]
        pivot_vals[partner][date] = int(valid.min()) if len(valid) else None
        out = group["StockAvailability"].fillna("").str.lower()
        stock_vals[partner][date] = bool(out.str.contains(r"out|sold").any())

    changes = {}
    for partner in pivot_vals:
        changes[partner] = {}
        prev = pivot_vals[partner].get(prev_date) if prev_date else None
        for date in dates:
            cur = pivot_vals[partner].get(date)
            changes[partner][date] = (prev is not None and cur is not None and cur != prev)
            if cur is not None:
                prev = cur

    for partner in list(pivot_vals.keys()):
        pivot_vals[partner] = {d: v for d, v in pivot_vals[partner].items() if d in dates}
        stock_vals[partner] = {d: v for d, v in stock_vals[partner].items() if d in dates}

    latest_date = all_dates[-1] if all_dates else None
    best_prices = {}
    if latest_date:
        df_latest = df[df["Date"] == latest_date]
        for col, label in [
            ("Lists_Price", "List Price"), ("Selling_Price", "Selling Price"),
            ("Pointback", "Pointback"), ("PriceMinusPoints", "Price - Points"),
        ]:
            valid = df_latest[(df_latest[col] > 0) & (df_latest["Partner"] != "-")]
            if not valid.empty:
                idx = valid[col].idxmin()
                best_prices[label] = {
                    "partner": str(valid.loc[idx, "Partner"]),
                    "value":   int(valid.loc[idx, col]),
                    "date":    latest_date,
                }

    return jsonify({
        "dates":    dates,
        "partners": partners,
        "values":   pivot_vals,
        "stock":    stock_vals,
        "changes":  changes,
        "param":    param,
        "best_prices": best_prices,
    })


@app.route("/api/comments")
@require_login
def api_comments():
    return jsonify(COMMENTS)


@app.route("/api/comment", methods=["POST"])
@require_login
def api_comment_post():
    data = request.get_json(force=True)
    save_comment(data.get("partner",""), data.get("date",""),
                 data.get("price",0), data.get("comment",""))
    return jsonify({"ok": True})


@app.route("/api/comment", methods=["DELETE"])
@require_login
def api_comment_delete():
    data = request.get_json(force=True)
    delete_comment(data.get("partner",""), data.get("date",""))
    return jsonify({"ok": True})


# ── HTML テンプレート ───────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CES Dashboard</title>
<style>
  :root {
    --bg: #f4f6f9; --header-bg: #141829; --bar-bg: #1e2336;
    --accent: #4f8ef7; --accent2: #34c994; --text: #2c3e50; --muted: #8a93a8;
    --card: #ffffff; --border: #e0e5ef; --danger: #e74c3c; --warn: #f39c12;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         background: var(--bg); color: var(--text);
         display: flex; flex-direction: column; height: 100vh; overflow: hidden; }

  /* ── Header ── */
  #header {
    background: var(--header-bg); color: #fff; padding: 11px 20px;
    display: flex; align-items: center; gap: 14px; flex-shrink: 0;
  }
  #header h2 { font-size: 17px; font-weight: 700; flex: 1; }
  #header h2 span { color: var(--accent); }
  #row-count { font-size: 12px; color: var(--muted); white-space: nowrap; }
  .param-tabs { display: flex; gap: 4px; flex-wrap: nowrap; }
  .param-btn {
    padding: 4px 11px; border-radius: 20px; border: 1px solid #3a4468;
    background: #2c3450; color: #c8d0e0; cursor: pointer; font-size: 11px;
    transition: all 0.15s; white-space: nowrap;
  }
  .param-btn.active { background: var(--accent); border-color: var(--accent); color: #fff; }
  .param-btn:hover:not(.active) { background: #3a4468; }
  .logout-btn {
    padding: 4px 12px; border-radius: 20px; border: 1px solid #3a4468;
    background: transparent; color: #8a93a8; cursor: pointer; font-size: 11px;
    transition: all 0.15s; white-space: nowrap; text-decoration: none;
  }
  .logout-btn:hover { border-color: var(--danger); color: var(--danger); }
  .user-email { font-size: 11px; color: #6a7490; white-space: nowrap; }

  /* ── Filter bar ── */
  #filter-bar {
    background: var(--bar-bg); padding: 10px 18px 10px;
    display: flex; flex-direction: column; gap: 8px;
    flex-shrink: 0; border-bottom: 1px solid #0d1020;
  }
  #filter-row-top {
    display: flex; align-items: center; gap: 10px; flex-wrap: wrap;
  }

  /* Date pickers */
  .date-group { display: flex; align-items: center; gap: 6px; }
  .date-group span { font-size: 11px; color: #8a93a8; white-space: nowrap; }
  .date-input {
    background: #2c3450; border: 1px solid #3a4468; color: #c8d0e0;
    border-radius: 6px; padding: 5px 9px; font-size: 12px; width: 130px;
  }

  /* Filter chips */
  #filter-container { display: grid; grid-template-columns: repeat(9, 1fr); gap: 6px; align-items: start; }
  .filter-chip { position: relative; }
  .filter-chip.hidden { display: none; }

  .chip-btn {
    display: flex; align-items: center; justify-content: space-between; gap: 5px;
    width: 100%;
    background: #2c3450; border: 1px solid #3a4468; color: #c8d0e0;
    border-radius: 6px; padding: 6px 10px; font-size: 12px; cursor: pointer;
    transition: border-color 0.15s, background 0.15s;
    min-height: 32px; overflow: hidden;
  }
  .chip-btn:hover { border-color: var(--accent); background: #243058; }
  .chip-btn.active { border-color: var(--accent); background: #1a3060; }
  .chip-btn.has-selection { border-color: var(--accent); background: #1a3060; color: #fff; }
  .chip-label { font-weight: 500; letter-spacing: 0.2px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .chip-count { color: #7eb8ff; font-weight: 700; font-size: 11px; flex-shrink: 0; }
  .chip-arrow { font-size: 9px; opacity: 0.5; flex-shrink: 0; }

  .chip-dropdown {
    display: none; position: absolute; top: calc(100% + 5px); left: 0;
    background: #fff; border: 1px solid var(--border); border-radius: 8px;
    box-shadow: 0 8px 24px rgba(0,0,0,0.18); z-index: 200;
    min-width: 180px; max-height: 260px; overflow-y: auto;
  }
  .chip-dropdown.open { display: block; }
  .chip-option {
    display: flex; align-items: center; gap: 8px;
    padding: 7px 14px; font-size: 13px; cursor: pointer;
    color: var(--text); user-select: none; line-height: 1.3;
  }
  .chip-option:hover { background: #f0f4ff; }
  .chip-option:has(input:checked) { background: #e8f0fe; color: #1d4ed8; font-weight: 600; }
  .chip-option input[type=checkbox] {
    cursor: pointer; accent-color: var(--accent);
    width: 14px; height: 14px; flex-shrink: 0;
  }
  .chip-option.option-hidden { display: none; }

  /* Reset button */
  .reset-btn {
    padding: 6px 14px; border-radius: 6px; min-height: 32px;
    background: transparent; border: 1px solid #3a4468; color: #8a93a8;
    font-size: 12px; cursor: pointer; transition: all 0.15s; white-space: nowrap;
  }
  .reset-btn:hover { border-color: var(--danger); color: var(--danger); background: #2c1e1e; }

  /* BD Focus toggle */
  .bd-focus-btn {
    padding: 6px 14px; border-radius: 6px; min-height: 32px;
    background: transparent; border: 1px solid #3a4468; color: #8a93a8;
    font-size: 12px; font-weight: 600; cursor: pointer; transition: all 0.15s; white-space: nowrap;
  }
  .bd-focus-btn:hover:not(.active) { border-color: #f39c12; color: #f39c12; }
  .bd-focus-btn.active { background: #f39c12; border-color: #f39c12; color: #fff; }

  /* ── Content ── */
  #content { flex: 1; overflow-y: auto; padding: 18px 22px; }

  /* KPI cards */
  .kpi-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin-bottom: 18px; }
  .kpi-card {
    background: var(--card); border-radius: 10px; padding: 14px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06); border-left: 4px solid var(--accent);
  }
  .kpi-card.k2 { border-left-color: var(--accent2); }
  .kpi-card.k3 { border-left-color: var(--warn); }
  .kpi-card.k4 { border-left-color: var(--danger); }
  .kpi-label { font-size: 11px; color: var(--muted); font-weight: 600;
               text-transform: uppercase; letter-spacing: 0.7px; }
  .kpi-value { font-size: 24px; font-weight: 700; color: var(--text); margin-top: 4px; }
  .kpi-sub   { font-size: 11px; color: var(--muted); margin-top: 2px; }

  /* Charts row */
  .charts-row { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 18px; }
  .chart-card {
    background: var(--card); border-radius: 10px; padding: 14px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
  }
  .chart-card h3 { font-size: 12px; font-weight: 600; color: var(--text); margin-bottom: 10px; }
  .bar-row { display: flex; align-items: center; margin-bottom: 6px; }
  .bar-label { font-size: 11px; width: 120px; text-align: right; padding-right: 8px;
               color: var(--text); white-space: nowrap; overflow: hidden;
               text-overflow: ellipsis; flex-shrink: 0; }
  .bar-track { flex: 1; background: #f0f2f7; border-radius: 4px; height: 16px; }
  .bar-fill { height: 100%; border-radius: 4px; transition: width 0.4s ease; }
  .bar-val { font-size: 11px; color: var(--muted); margin-left: 6px; width: 70px; flex-shrink: 0; }

  /* Trend chart */
  .trend-card {
    background: var(--card); border-radius: 10px; padding: 14px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06); margin-bottom: 18px;
  }
  .trend-card h3 { font-size: 12px; font-weight: 600; margin-bottom: 8px; }

  /* Table */
  .table-card {
    background: var(--card); border-radius: 10px; padding: 14px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06); margin-bottom: 18px;
  }
  .table-header { display: flex; align-items: center; gap: 12px; flex-wrap: wrap; margin-bottom: 8px; }
  .table-header h3 { font-size: 12px; font-weight: 600; flex-shrink: 0; }

  /* Table multi-select filter chips */
  .tbl-chip { position: relative; }
  .tbl-chip-btn {
    display: flex; align-items: center; gap: 5px;
    background: #f7f9fc; border: 1px solid var(--border); color: var(--text);
    border-radius: 6px; padding: 5px 10px; font-size: 12px; cursor: pointer;
    white-space: nowrap; transition: border-color 0.15s, background 0.15s; min-height: 30px;
  }
  .tbl-chip-btn:hover { border-color: var(--accent); }
  .tbl-chip-btn.has-sel { border-color: var(--accent); background: #e8f0fe; color: #1d4ed8; font-weight: 600; }
  .tbl-chip-cnt { color: #1d4ed8; font-weight: 700; font-size: 11px; }
  .tbl-chip-arr { font-size: 9px; opacity: 0.45; }
  .tbl-chip-dd {
    display: none; position: absolute; top: calc(100% + 4px); left: 0;
    background: #fff; border: 1px solid var(--border); border-radius: 8px;
    box-shadow: 0 8px 24px rgba(0,0,0,0.14); z-index: 300;
    min-width: 190px; max-height: 240px; overflow-y: auto;
  }
  .tbl-chip-dd.open { display: block; }
  .tbl-chip-opt {
    display: flex; align-items: center; gap: 8px;
    padding: 7px 12px; font-size: 12px; cursor: pointer;
    color: var(--text); user-select: none;
  }
  .tbl-chip-opt:hover { background: #f0f4ff; }
  .tbl-chip-opt input[type=checkbox] {
    cursor: pointer; accent-color: var(--accent); width: 13px; height: 13px; flex-shrink: 0;
  }
  .table-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: 12px; }
  thead th {
    background: #f7f9fc; color: var(--muted); font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.5px; font-size: 10px;
    padding: 7px 10px; border-bottom: 2px solid var(--border);
    text-align: left; white-space: nowrap;
  }
  tbody td { padding: 6px 10px; border-bottom: 1px solid var(--border); white-space: nowrap; }
  tbody tr:hover { background: #f7f9fc; }
  .badge { display: inline-block; padding: 2px 7px; border-radius: 12px; font-size: 10px; font-weight: 600; }
  .badge-ok  { background: #d4f0e8; color: #1a7a54; }
  .badge-err { background: #fde8e8; color: #c0392b; }
  .badge-in  { background: #dbeafe; color: #1d4ed8; }
  .badge-out { background: #fee2e2; color: #b91c1c; }
  .badge-few { background: #fef3c7; color: #92400e; }

  .loading { text-align: center; padding: 40px; color: var(--muted); }
  .spinner { display: inline-block; width: 20px; height: 20px;
             border: 3px solid var(--border); border-top-color: var(--accent);
             border-radius: 50%; animation: spin 0.7s linear infinite; }
  @keyframes spin { to { transform: rotate(360deg); } }

  /* Pivot table */
  .pivot-card {
    background: var(--card); border-radius: 10px; padding: 14px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06); margin-bottom: 18px;
  }
  .pivot-card h3 { font-size: 12px; font-weight: 600; margin-bottom: 8px; }
  .pivot-wrap { overflow-x: auto; max-height: 340px; overflow-y: auto; }
  .pivot-table { border-collapse: collapse; font-size: 11px; white-space: nowrap; }
  .pivot-table thead th {
    position: sticky; top: 0; z-index: 2;
    background: #f7f9fc; color: var(--muted); font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.4px; font-size: 10px;
    padding: 6px 10px; border-bottom: 2px solid var(--border);
    border-right: 1px solid var(--border);
  }
  .pivot-table thead th.partner-head {
    position: sticky; left: 0; z-index: 3; min-width: 100px; text-align: left;
  }
  .pivot-table thead th.date-head { text-align: center; min-width: 74px; }
  .pivot-table tbody td {
    padding: 5px 10px; border-bottom: 1px solid var(--border);
    border-right: 1px solid var(--border); text-align: right; background: #fff;
  }
  .pivot-table tbody td.partner-cell {
    position: sticky; left: 0; z-index: 1;
    background: #f7f9fc; font-weight: 600; text-align: left;
    border-right: 2px solid var(--border);
  }
  .pivot-table tbody tr:hover td { filter: brightness(0.97); }
  .cell-changed { color: #1d4ed8; font-weight: 700; }
  .cell-null { color: #ccc; }
  .cell-cartoff { display: block; font-size: 9px; color: #e74c3c; font-weight: 600;
                  margin-top: 2px; letter-spacing: 0.3px; }

  /* Partner line chart */
  .partner-chart-card {
    background: var(--card); border-radius: 10px; padding: 14px 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06); margin-bottom: 18px;
  }
  .partner-chart-card h3 { font-size: 12px; font-weight: 600; margin-bottom: 6px; }
  #partnerChartCanvas { display: block; width: 100%; cursor: pointer; }
  .chart-legend { display: flex; flex-wrap: wrap; gap: 10px; margin-top: 8px; }
  .legend-item { display: flex; align-items: center; gap: 5px; font-size: 11px; color: var(--text); }
  .legend-dot { width: 11px; height: 11px; border-radius: 50%; flex-shrink: 0; }

  /* Best prices row */
  .best-prices-row {
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-top: 14px;
  }
  .best-price-card {
    background: #f7f9fc; border-radius: 8px; padding: 10px 14px;
    border-left: 3px solid var(--accent2);
  }
  .best-price-label { font-size: 10px; color: var(--muted); font-weight: 600;
                      text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px; }
  .best-price-partner { font-size: 13px; font-weight: 700; color: var(--text); }
  .best-price-value { font-size: 16px; font-weight: 700; color: var(--accent2); margin-top: 2px; }
  .best-price-date { font-size: 10px; color: var(--muted); margin-top: 2px; }

  /* Comment modal */
  #comment-modal-overlay {
    display: none; position: fixed; inset: 0;
    background: rgba(0,0,0,0.45); z-index: 1000;
    align-items: center; justify-content: center;
  }
  #comment-modal-overlay.open { display: flex; }
  #comment-modal {
    background: #fff; border-radius: 12px; padding: 24px 26px;
    width: 380px; box-shadow: 0 16px 48px rgba(0,0,0,0.28);
  }
  #comment-modal h4 { font-size: 14px; font-weight: 700; color: var(--text); margin-bottom: 4px; }
  #comment-modal .cm-meta { font-size: 12px; color: var(--muted); margin-bottom: 14px; }
  #comment-modal textarea {
    width: 100%; height: 90px; border: 1px solid var(--border); border-radius: 8px;
    padding: 10px 12px; font-size: 13px; resize: vertical; font-family: inherit;
    outline: none; transition: border-color 0.15s; color: var(--text);
  }
  #comment-modal textarea:focus { border-color: var(--accent); }
  .cm-actions { display: flex; gap: 8px; margin-top: 12px; }
  .cm-delete {
    padding: 8px 14px; border-radius: 6px; border: 1px solid var(--danger);
    background: #fff; color: var(--danger); cursor: pointer; font-size: 13px;
    transition: all 0.15s; margin-right: auto;
  }
  .cm-delete:hover { background: var(--danger); color: #fff; }
  .cm-cancel {
    padding: 8px 18px; border-radius: 6px; border: 1px solid var(--border);
    background: #fff; color: var(--muted); cursor: pointer; font-size: 13px;
    transition: all 0.15s;
  }
  .cm-cancel:hover { border-color: var(--danger); color: var(--danger); }
  .cm-save {
    padding: 8px 20px; border-radius: 6px; border: none;
    background: var(--accent); color: #fff; cursor: pointer; font-size: 13px; font-weight: 600;
    transition: background 0.15s;
  }
  .cm-save:hover { background: #3a7ae0; }
</style>
</head>
<body>

<!-- Header -->
<div id="header">
  <h2>CES <span>Dashboard</span></h2>
  <span id="row-count">Loading…</span>
  <div class="param-tabs" id="param-tabs"></div>
  <span class="user-email" id="user-email">USER_EMAIL_PLACEHOLDER</span>
  <a href="/logout" class="logout-btn">Logout</a>
</div>

<!-- Filter bar -->
<div id="filter-bar">
  <div id="filter-row-top">
    <div class="date-group">
      <span>From</span>
      <input type="date" class="date-input" id="f_date_from" value="DATE_15_PLACEHOLDER">
      <span>To</span>
      <input type="date" class="date-input" id="f_date_to" value="DATE_MAX_PLACEHOLDER">
    </div>
    <button class="bd-focus-btn" id="bd-focus-btn" onclick="toggleBdFocus()">BD Focus</button>
    <button class="reset-btn" onclick="clearFilters()">Reset Filters</button>
  </div>
  <div id="filter-container"></div>
</div>

<!-- Content -->
<div id="content">
  <div class="loading"><div class="spinner"></div></div>
</div>

<!-- Comment Modal -->
<div id="comment-modal-overlay">
  <div id="comment-modal">
    <h4>Comment</h4>
    <div class="cm-meta" id="cm-meta"></div>
    <textarea id="cm-text" placeholder="Enter comment…"></textarea>
    <div class="cm-actions">
      <button class="cm-delete" id="cm-delete-btn" onclick="deleteComment()" style="display:none">Delete</button>
      <button class="cm-cancel" onclick="closeCommentModal()">Cancel</button>
      <button class="cm-save" onclick="submitComment()">Save</button>
    </div>
  </div>
</div>

<script>
const VALUE_COLS = ["PriceMinusPoints","Selling_Price","Lists_Price",
                    "Pointback","Hidden_Discount","Displayed_Discount","Total_Discount"];
const VALUE_LABELS = {
  PriceMinusPoints:   "Price - Points",
  Selling_Price:      "Selling Price",
  Lists_Price:        "List Price",
  Pointback:          "Pointback",
  Hidden_Discount:    "Hidden Discount",
  Displayed_Discount: "Displayed Discount",
  Total_Discount:     "Total Discount",
};
const CHART_COLORS = [
  "#4f8ef7","#34c994","#f39c12","#e74c3c","#9b59b6",
  "#1abc9c","#e67e22","#3498db","#2ecc71","#e91e63",
];
const PARTNER_COLORS = [
  "#4f8ef7","#e74c3c","#34c994","#f39c12","#9b59b6",
  "#1abc9c","#e67e22","#3498db","#e91e63","#00bcd4",
];

let currentParam = "Selling_Price";
let debounceTimer = null;
let options = {};
let _tableStore = { rows: [], paramCol: "Selling_Price", tableMode: "latest" };
let _comments = {};
let _commentTarget = null;
let bdFocusOn = false;
const DATE_MAX = "DATE_MAX_JS_PLACEHOLDER";
const DATE_MIN = "DATE_MIN_JS_PLACEHOLDER";
const DATE_15  = "DATE_15_JS_PLACEHOLDER";

// ── Init ──────────────────────────────────────────────────────────────────────
async function init() {
  const resp = await fetch("/api/options");
  options = await resp.json();
  buildParamTabs();
  buildFilters();
  await loadComments();
  refresh();
}

// ── Comments ──────────────────────────────────────────────────────────────────
async function loadComments() {
  try {
    const data = await (await fetch("/api/comments")).json();
    _comments = {};
    data.forEach(c => {
      if (c.Partner && c.Date) _comments[c.Partner + "|" + c.Date] = c.Comment || "";
    });
  } catch(e) {}
}

function openCommentModal(partner, date, price) {
  _commentTarget = { partner, date, price };
  document.getElementById("cm-meta").textContent =
    partner + "  ·  " + date + "  ·  ¥" + Math.round(price).toLocaleString("ja-JP");
  const key = partner + "|" + date;
  const existing = _comments[key] || "";
  document.getElementById("cm-text").value = existing;
  document.getElementById("cm-delete-btn").style.display = existing ? "inline-block" : "none";
  document.getElementById("comment-modal-overlay").classList.add("open");
  setTimeout(() => document.getElementById("cm-text").focus(), 60);
}

function closeCommentModal() {
  document.getElementById("comment-modal-overlay").classList.remove("open");
  _commentTarget = null;
}

async function submitComment() {
  if (!_commentTarget) return;
  const text = document.getElementById("cm-text").value.trim();
  const { partner, date, price } = _commentTarget;
  const key = partner + "|" + date;
  _comments[key] = text;
  closeCommentModal();
  try {
    await fetch("/api/comment", {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify({ partner, date, price, comment: text }),
    });
  } catch(e) {}
  if (_partnerChartState) drawPartnerChart(_partnerChartState, null);
}

async function deleteComment() {
  if (!_commentTarget) return;
  const { partner, date } = _commentTarget;
  const key = partner + "|" + date;
  delete _comments[key];
  closeCommentModal();
  try {
    await fetch("/api/comment", {
      method: "DELETE",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify({ partner, date }),
    });
  } catch(e) {}
  if (_partnerChartState) drawPartnerChart(_partnerChartState, null);
}

document.addEventListener("DOMContentLoaded", () => {
  document.getElementById("comment-modal-overlay").addEventListener("click", function(e) {
    if (e.target === this) closeCommentModal();
  });
});

function buildParamTabs() {
  const tabs = document.getElementById("param-tabs");
  VALUE_COLS.forEach(col => {
    const btn = document.createElement("button");
    btn.className = "param-btn" + (col === currentParam ? " active" : "");
    btn.textContent = VALUE_LABELS[col];
    btn.onclick = () => { currentParam = col; updateParamTabs(); refresh(); };
    btn.id = "tab_" + col;
    tabs.appendChild(btn);
  });
}
function updateParamTabs() {
  VALUE_COLS.forEach(col =>
    document.getElementById("tab_" + col).classList.toggle("active", col === currentParam));
}

// ── Filter chips ──────────────────────────────────────────────────────────────
function buildFilters() {
  const container = document.getElementById("filter-container");
  const FCOLS = Object.keys(options).filter(k => k !== "Date");
  FCOLS.forEach(col => {
    const chip = document.createElement("div");
    chip.className = "filter-chip";
    chip.id = "fc_" + col;

    const btn = document.createElement("button");
    btn.className = "chip-btn";
    btn.id = "cb_" + col;
    btn.innerHTML =
      `<span class="chip-label">${col.replace(/_/g, " ")}</span>` +
      `<span class="chip-count" id="cc_${col}"></span>` +
      `<span class="chip-arrow">▾</span>`;
    btn.addEventListener("click", e => { e.stopPropagation(); toggleDropdown(col); });

    const dd = document.createElement("div");
    dd.className = "chip-dropdown";
    dd.id = "cd_" + col;
    dd.addEventListener("click", e => e.stopPropagation());

    const opts = document.createElement("div");
    opts.id = "co_" + col;
    options[col].forEach(val => {
      const lbl = document.createElement("label");
      lbl.className = "chip-option";
      const cb = document.createElement("input");
      cb.type = "checkbox"; cb.value = val;
      cb.addEventListener("change", () => { updateChipCount(col); scheduleRefresh(); });
      lbl.appendChild(cb);
      lbl.appendChild(document.createTextNode(" " + val));
      opts.appendChild(lbl);
    });

    dd.appendChild(opts);
    chip.appendChild(btn);
    chip.appendChild(dd);
    container.appendChild(chip);
  });

  document.addEventListener("click", closeAllDropdowns);
  document.getElementById("f_date_from").addEventListener("change", scheduleRefresh);
  document.getElementById("f_date_to").addEventListener("change", scheduleRefresh);
}

function toggleDropdown(col) {
  const dd = document.getElementById("cd_" + col);
  const wasOpen = dd.classList.contains("open");
  closeAllDropdowns();
  if (!wasOpen) {
    dd.classList.add("open");
    document.getElementById("cb_" + col).classList.add("active");
  }
}
function closeAllDropdowns() {
  document.querySelectorAll(".chip-dropdown.open").forEach(d => {
    d.classList.remove("open");
    const col = d.id.replace("cd_", "");
    const btn = document.getElementById("cb_" + col);
    if (btn) btn.classList.remove("active");
  });
}
function updateChipCount(col) {
  const n = document.querySelectorAll(`#co_${col} input:checked`).length;
  const el = document.getElementById("cc_" + col);
  if (el) el.textContent = n > 0 ? ` (${n})` : "";
}

function scheduleRefresh() {
  clearTimeout(debounceTimer);
  debounceTimer = setTimeout(refresh, 300);
}
function toggleBdFocus() {
  bdFocusOn = !bdFocusOn;
  document.getElementById("bd-focus-btn").classList.toggle("active", bdFocusOn);
  refresh();
}
function clearFilters() {
  document.querySelectorAll(".chip-options input, #filter-container input[type=checkbox]")
    .forEach(cb => cb.checked = false);
  document.querySelectorAll(".chip-count").forEach(el => el.textContent = "");
  Object.keys(options).filter(k => k !== "Date").forEach(col => {
    document.querySelectorAll(`#co_${col} input`).forEach(cb => cb.checked = false);
    updateChipCount(col);
  });
  document.getElementById("f_date_from").value = DATE_15;
  document.getElementById("f_date_to").value = DATE_MAX;
  bdFocusOn = false;
  document.getElementById("bd-focus-btn").classList.remove("active");
  refresh();
}

// ── Build query ───────────────────────────────────────────────────────────────
function buildQuery() {
  const parts = ["param=" + currentParam];
  parts.push("date_from=" + document.getElementById("f_date_from").value);
  parts.push("date_to="   + document.getElementById("f_date_to").value);
  if (bdFocusOn) parts.push("bd_focus=true");
  Object.keys(options).filter(k => k !== "Date").forEach(col => {
    document.querySelectorAll(`#co_${col} input:checked`).forEach(cb =>
      parts.push(encodeURIComponent(col) + "=" + encodeURIComponent(cb.value)));
  });
  return parts.join("&");
}

// ── Refresh ───────────────────────────────────────────────────────────────────
async function refresh() {
  const query = buildQuery();
  const resp = await fetch("/api/data?" + query);
  const data = await resp.json();
  render(data);
}

// ── Cascade filter update ─────────────────────────────────────────────────────
function updateFilterOptions(avail) {
  Object.keys(options).filter(k => k !== "Date").forEach(col => {
    const chip = document.getElementById("fc_" + col);
    if (!chip) return;
    const availSet  = new Set(avail[col] || []);
    const anySelected = document.querySelectorAll(`#co_${col} input:checked`).length > 0;
    const meaningful  = (avail[col] || []).filter(v => v !== "-");

    if (meaningful.length <= 1 && !anySelected) {
      chip.classList.add("hidden"); return;
    }
    chip.classList.remove("hidden");
    document.querySelectorAll(`#co_${col} label.chip-option`).forEach(lbl => {
      const cb = lbl.querySelector("input");
      if (anySelected) {
        lbl.classList.remove("option-hidden");
        lbl.style.opacity = "1";
      } else {
        const inAvail = availSet.has(cb.value);
        if (!inAvail && !cb.checked) {
          lbl.classList.add("option-hidden");
        } else {
          lbl.classList.remove("option-hidden");
          lbl.style.opacity = inAvail ? "1" : "0.5";
        }
      }
    });
  });
}

// ── Render ────────────────────────────────────────────────────────────────────
function fmt(n) {
  if (n === null || n === undefined) return "-";
  return n.toLocaleString("ja-JP");
}

function render(data) {
  document.getElementById("row-count").textContent =
    `${data.total_rows.toLocaleString()} rows`;

  document.getElementById("content").innerHTML = `
    <!-- Pivot table -->
    <div class="pivot-card">
      <h3 id="pivot-title">Partner × Date Pivot Table</h3>
      <div class="pivot-wrap" id="pivot-table-wrap">
        <div class="loading"><div class="spinner"></div></div>
      </div>
    </div>
    <!-- Partner line chart -->
    <div class="partner-chart-card">
      <h3 id="partner-chart-title">Partner Price Trend</h3>
      <canvas id="partnerChartCanvas"></canvas>
      <div class="chart-legend" id="partnerLegend"></div>
      <div class="best-prices-row" id="bestPricesRow"></div>
    </div>
    <!-- Table -->
    <div class="table-card">
      <div class="table-header">
        <h3 id="table-title">Data Table (latest date)</h3>
        <div class="tbl-chip" id="tbl-chip-partner">
          <button class="tbl-chip-btn" id="tbl-btn-partner" onclick="toggleTblDD('partner')">
            Partner <span class="tbl-chip-cnt" id="tbl-cnt-partner"></span><span class="tbl-chip-arr">▾</span>
          </button>
          <div class="tbl-chip-dd" id="tbl-dd-partner"><div id="tbl-opts-partner"></div></div>
        </div>
        <div class="tbl-chip" id="tbl-chip-product">
          <button class="tbl-chip-btn" id="tbl-btn-product" onclick="toggleTblDD('product')">
            Product <span class="tbl-chip-cnt" id="tbl-cnt-product"></span><span class="tbl-chip-arr">▾</span>
          </button>
          <div class="tbl-chip-dd" id="tbl-dd-product"><div id="tbl-opts-product"></div></div>
        </div>
      </div>
      <div class="table-wrap" id="data-table"></div>
    </div>
  `;

  renderTable(data.table, data.kpis.param, data.table_mode);
  updateFilterOptions(data.available_options || {});
  loadPivot();
}

function stockBadge(v) {
  if (!v||v==="-") return `<span class="badge">${v}</span>`;
  const lv=v.toLowerCase();
  if (lv.includes("out")||lv.includes("sold")) return `<span class="badge badge-out">${v}</span>`;
  if (lv.includes("few"))   return `<span class="badge badge-few">${v}</span>`;
  if (lv.includes("in stock")) return `<span class="badge badge-in">${v}</span>`;
  return `<span class="badge">${v}</span>`;
}
function robotBadge(v) {
  return v==="OK"?`<span class="badge badge-ok">OK</span>`:`<span class="badge badge-err">${v}</span>`;
}

// ── Table multi-select filter helpers ─────────────────────────────────────────
function populateTblFilter(key, values, keepSel) {
  const container = document.getElementById(`tbl-opts-${key}`);
  if (!container) return;
  container.innerHTML = values.map(v =>
    `<label class="tbl-chip-opt">` +
    `<input type="checkbox" value="${v}"${keepSel.has(v) ? " checked" : ""}` +
    ` onchange="onTblFilterChange()"> ${v}</label>`
  ).join("");
  updateTblCount(key);
}

function updateTblCount(key) {
  const n = document.querySelectorAll(`#tbl-opts-${key} input:checked`).length;
  const cnt = document.getElementById(`tbl-cnt-${key}`);
  const btn = document.getElementById(`tbl-btn-${key}`);
  if (cnt) cnt.textContent = n > 0 ? `(${n})` : "";
  if (btn) btn.classList.toggle("has-sel", n > 0);
}

function toggleTblDD(key) {
  const dd = document.getElementById(`tbl-dd-${key}`);
  const wasOpen = dd.classList.contains("open");
  closeTblDDs();
  if (!wasOpen) dd.classList.add("open");
}
function closeTblDDs() {
  document.querySelectorAll(".tbl-chip-dd.open").forEach(d => d.classList.remove("open"));
}
document.addEventListener("click", e => {
  if (!e.target.closest(".tbl-chip")) closeTblDDs();
});

function renderTable(rows, paramCol, tableMode) {
  _tableStore = { rows, paramCol, tableMode };
  const partners = [...new Set(rows.map(r => r["Partner"]).filter(p => p && p !== "-"))].sort();
  const products = [...new Set(rows.map(r => r["Product"]).filter(p => p && p !== "-"))].sort();
  populateTblFilter("partner", partners, new Set());
  populateTblFilter("product", products, new Set());
  _drawTable(rows, paramCol, tableMode);
}

function onTblFilterChange() {
  updateTblCount("partner");
  updateTblCount("product");
  filterTable();
}

async function filterTable() {
  const pvals = [...document.querySelectorAll("#tbl-opts-partner input:checked")].map(c => c.value);
  const vals  = [...document.querySelectorAll("#tbl-opts-product input:checked")].map(c => c.value);
  const { rows, paramCol, tableMode } = _tableStore;

  if (pvals.length > 0 && tableMode === "latest") {
    const parts = ["param=" + currentParam];
    parts.push("date_from=" + document.getElementById("f_date_from").value);
    parts.push("date_to="   + document.getElementById("f_date_to").value);
    Object.keys(options).filter(k => k !== "Date" && k !== "Partner").forEach(col => {
      document.querySelectorAll(`#co_${col} input:checked`).forEach(cb =>
        parts.push(encodeURIComponent(col) + "=" + encodeURIComponent(cb.value)));
    });
    pvals.forEach(p => parts.push("Partner=" + encodeURIComponent(p)));

    const data = await (await fetch("/api/data?" + parts.join("&"))).json();
    const prevProdSel = new Set(vals);
    const products = [...new Set(data.table.map(r => r["Product"]).filter(p => p && p !== "-"))].sort();
    populateTblFilter("product", products, prevProdSel);
    let fresh = data.table;
    if (vals.length > 0) fresh = fresh.filter(r => vals.includes(r["Product"]));
    _drawTable(fresh, paramCol, "per_product");

  } else if (pvals.length === 0 && tableMode === "latest") {
    const prevProdSel = new Set(vals);
    const products = [...new Set(rows.map(r => r["Product"]).filter(p => p && p !== "-"))].sort();
    populateTblFilter("product", products, prevProdSel);
    const filtered = vals.length > 0 ? rows.filter(r => vals.includes(r["Product"])) : rows;
    _drawTable(filtered, paramCol, tableMode);

  } else {
    let filtered = rows;
    if (pvals.length > 0) filtered = filtered.filter(r => pvals.includes(r["Partner"]));
    if (vals.length > 0)  filtered = filtered.filter(r => vals.includes(r["Product"]));
    _drawTable(filtered, paramCol, tableMode);
  }
}

function _drawTable(rows, paramCol, tableMode) {
  const el = document.getElementById("data-table");
  const titleEl = document.getElementById("table-title");
  if (!rows.length) { el.innerHTML="<p style='color:#aaa;padding:16px'>No data</p>"; return; }

  let cols, labels;
  if (tableMode === "per_product") {
    if (titleEl) titleEl.textContent = "Latest Data per Product";
    const uniqPartners = new Set(rows.map(r => r["Partner"]).filter(p => p && p !== "-"));
    const showPartner = uniqPartners.size > 1;
    cols = [...(showPartner ? ["Partner"] : []),
            "Product","StockAvailability","Robot_Status",
            "Selling_Price","Lists_Price","Pointback",
            "Hidden_Discount","Displayed_Discount","Total_Discount","PriceMinusPoints"];
    labels = {
      Partner:"Partner", Product:"Product", StockAvailability:"Stock", Robot_Status:"Status",
      Selling_Price:"Selling Price", Lists_Price:"List Price",
      Pointback:"Pointback", Hidden_Discount:"Hidden Disc",
      Displayed_Discount:"Disp Disc", Total_Discount:"Total Disc",
      PriceMinusPoints:"Price - Points"
    };
  } else {
    if (titleEl) titleEl.textContent = "Data Table (latest date)";
    cols = ["Date","Partner","Product","StockAvailability","Robot_Status",
            "Selling_Price","Lists_Price","Pointback",
            "Hidden_Discount","Displayed_Discount","Total_Discount","PriceMinusPoints"];
    labels = {
      Date:"Date", Partner:"Partner", Product:"Product",
      StockAvailability:"Stock", Robot_Status:"Status",
      Selling_Price:"Selling Price", Lists_Price:"List Price",
      Pointback:"Pointback", Hidden_Discount:"Hidden Disc",
      Displayed_Discount:"Disp Disc", Total_Discount:"Total Disc",
      PriceMinusPoints:"Price - Points"
    };
  }

  const head = cols.map(c =>
    `<th${c===paramCol?" style='background:#e8f0fe;color:#1d4ed8'":""}>${labels[c]||c}</th>`
  ).join("");
  const body = rows.map(r => {
    const cells = cols.map(c => {
      let v = r[c] ?? "–";
      if (c === "StockAvailability") return `<td>${stockBadge(v)}</td>`;
      if (c === "Robot_Status")      return `<td>${robotBadge(v)}</td>`;
      const hl = c === paramCol ? " style='background:#f0f4ff;font-weight:600'" : "";
      if (typeof v === "number") return `<td${hl}>¥${fmt(v)}</td>`;
      return `<td${hl}>${v}</td>`;
    }).join("");
    return `<tr>${cells}</tr>`;
  }).join("");
  el.innerHTML = `<table><thead><tr>${head}</tr></thead><tbody>${body}</tbody></table>`;
}

// ── Pivot ─────────────────────────────────────────────────────────────────────
async function loadPivot() {
  const wrap = document.getElementById("pivot-table-wrap");
  if (!wrap) return;
  wrap.innerHTML = '<div class="loading"><div class="spinner"></div></div>';
  const data = await (await fetch("/api/pivot?" + buildQuery())).json();
  renderPivotTable(data);
  renderPartnerLineChart(data);
}

function renderPivotTable(data) {
  const wrap = document.getElementById("pivot-table-wrap");
  if (!wrap) return;
  const { dates, partners, values, stock, changes, param } = data;
  document.getElementById("pivot-title").textContent =
    `Partner × Date Pivot Table — ${VALUE_LABELS[param]||param} (Min, latest 15 days)`;
  if (!dates.length||!partners.length) { wrap.innerHTML="<p style='color:#aaa;padding:12px'>No data</p>"; return; }
  const head = `<tr><th class="partner-head">Partner</th>${dates.map(d=>`<th class="date-head">${d.slice(5)}</th>`).join("")}</tr>`;
  const body = partners.map(partner => {
    const cells = dates.map(date => {
      const val = values[partner] && values[partner][date];
      const chg = changes[partner] && changes[partner][date];
      const oos = stock[partner]  && stock[partner][date];
      const nullCls = (val === null || val === undefined) ? " cell-null" : "";
      const chgCls  = chg ? " cell-changed" : "";
      const disp    = (val !== null && val !== undefined) ? "¥" + val.toLocaleString() : "-";
      const cartTag = oos ? `<span class="cell-cartoff">Cart off</span>` : "";
      return `<td class="${nullCls}${chgCls}">${disp}${cartTag}</td>`;
    }).join("");
    return `<tr><td class="partner-cell">${partner}</td>${cells}</tr>`;
  }).join("");
  wrap.innerHTML=`<table class="pivot-table"><thead>${head}</thead><tbody>${body}</tbody></table>`;
}

// ── Partner line chart ────────────────────────────────────────────────────────
function fmtShort(v) {
  if (v >= 10000) return "¥" + Math.round(v/1000) + "K";
  return "¥" + Math.round(v).toLocaleString();
}

let _partnerChartState = null;

function renderPartnerLineChart(data) {
  const canvas = document.getElementById("partnerChartCanvas");
  const legend = document.getElementById("partnerLegend");
  if (!canvas || !legend) return;
  const { dates, partners, values, param } = data;
  document.getElementById("partner-chart-title").textContent =
    `Partner Price Trend — ${VALUE_LABELS[param]||param}`;
  legend.innerHTML = "";
  if (!dates.length || !partners.length) { canvas.style.display = "none"; return; }
  canvas.style.display = "block";

  const dpr = window.devicePixelRatio || 1;
  const W = Math.max(canvas.parentElement.getBoundingClientRect().width - 32, 400);
  const H = 340;
  canvas.width = W * dpr; canvas.height = H * dpr;
  canvas.style.width = W + 'px'; canvas.style.height = H + 'px';

  const padL=72, padR=24, padT=24, padB=54;
  const cW=W-padL-padR, cH=H-padT-padB;

  let allVals = [];
  partners.forEach(p => dates.forEach(d => {
    const v = values[p] && values[p][d]; if (v > 0) allVals.push(v);
  }));
  if (!allVals.length) {
    const ctx = canvas.getContext("2d");
    ctx.fillStyle = "#aaa"; ctx.font = "13px sans-serif"; ctx.textAlign = "center";
    ctx.fillText("No data", W/2, H/2); return;
  }
  const maxV = Math.max(...allVals), minV = Math.min(...allVals);
  const rng = maxV - minV || 1;
  const vMax = maxV + rng*0.08, vMin = Math.max(0, minV - rng*0.08);

  const allPoints = [];
  partners.forEach((partner, pi) => {
    const color = PARTNER_COLORS[pi % PARTNER_COLORS.length];
    const pts = dates.map((d, i) => {
      const v = values[partner] && values[partner][d];
      if (!v || v <= 0) return null;
      const x = padL + (dates.length > 1 ? i / (dates.length - 1) : 0.5) * cW;
      const y = padT + (1 - (v - vMin) / (vMax - vMin)) * cH;
      return { x, y, v, color, pi, partner, di: i };
    });
    allPoints.push({ partner, pi, color, pts });
  });

  const LW=44, LH=10, GAP=2;
  const candidates = [];
  allPoints.forEach(({ color, pts, pi, partner }) => {
    pts.forEach(p => { if (p) candidates.push({ x:p.x, y:p.y, v:p.v, color, pi, partner, di:p.di }); });
  });
  const tryOffsets = [-13, 13, -24, 24, -35, 35, -46, 46, -57, 57, -68, 68];
  const placed = [];
  const byX = {};
  candidates.forEach(c => {
    const key = Math.round(c.x);
    if (!byX[key]) byX[key] = [];
    byX[key].push(c);
  });
  Object.values(byX).forEach(group => {
    group.sort((a, b) => a.y - b.y);
    group.forEach(c => {
      const txt = fmtShort(c.v);
      const tw = Math.min(txt.length * 6.5, LW);
      const lx1 = c.x - tw/2 - 1, lx2 = c.x + tw/2 + 1;
      let placedY = c.y - 13;
      for (const off of tryOffsets) {
        const ty = c.y + off;
        if (ty < padT || ty > padT + cH + LH) continue;
        const ly1 = ty - LH, ly2 = ty + GAP;
        const collision = placed.some(p => lx1 < p.x2 && lx2 > p.x1 && ly1 < p.y2 && ly2 > p.y1);
        if (!collision) { placedY = ty; break; }
      }
      c.placedY = placedY;
      placed.push({ x1: lx1, y1: placedY - LH, x2: lx2, y2: placedY + GAP });
    });
  });

  _partnerChartState = { canvas, dates, allPoints, candidates, padL, padR, padT, padB, cW, cH, W, H, vMax, vMin, dpr };
  drawPartnerChart(_partnerChartState, null);

  canvas.onmousemove = function(e) {
    if (!_partnerChartState) return;
    const rect = canvas.getBoundingClientRect();
    const mx = e.clientX - rect.left;
    const my = e.clientY - rect.top;
    let nearest = null, minDist = 28;
    _partnerChartState.allPoints.forEach(({ pts, partner, color }) => {
      pts.forEach(p => {
        if (!p) return;
        const d = Math.hypot(p.x - mx, p.y - my);
        if (d < minDist) { minDist = d; nearest = { ...p, partner, color, date: _partnerChartState.dates[p.di] }; }
      });
    });
    drawPartnerChart(_partnerChartState, nearest);
  };
  canvas.onmouseleave = function() {
    if (_partnerChartState) drawPartnerChart(_partnerChartState, null);
  };

  canvas.onclick = function(e) {
    if (!_partnerChartState) return;
    const rect = canvas.getBoundingClientRect();
    const mx = e.clientX - rect.left;
    const my = e.clientY - rect.top;
    let nearest = null, minDist = 20;
    _partnerChartState.allPoints.forEach(({ pts, partner, color }) => {
      pts.forEach(p => {
        if (!p) return;
        const d = Math.hypot(p.x - mx, p.y - my);
        if (d < minDist) { minDist = d; nearest = { ...p, partner, color, date: _partnerChartState.dates[p.di] }; }
      });
    });
    if (nearest) openCommentModal(nearest.partner, nearest.date, nearest.v);
  };

  allPoints.forEach(({ partner, color }) => {
    const item = document.createElement("div");
    item.className = "legend-item";
    item.innerHTML = `<div class="legend-dot" style="background:${color}"></div>${partner}`;
    legend.appendChild(item);
  });

  const bpRow = document.getElementById("bestPricesRow");
  if (bpRow) {
    const bp = data.best_prices || {};
    const labels = ["List Price", "Selling Price", "Pointback", "Price - Points"];
    bpRow.innerHTML = labels.map(label => {
      const entry = bp[label];
      if (!entry) return `<div class="best-price-card"><div class="best-price-label">${label}</div><div class="best-price-partner" style="color:#ccc">—</div></div>`;
      return `<div class="best-price-card">
        <div class="best-price-label">Best ${label}</div>
        <div class="best-price-partner">${entry.partner}</div>
        <div class="best-price-value">¥${entry.value.toLocaleString("ja-JP")}</div>
        <div class="best-price-date">${entry.date}</div>
      </div>`;
    }).join("");
  }
}

function drawPartnerChart(state, hoverPt) {
  const { canvas, dates, allPoints, candidates, padL, padR, padT, padB, cW, cH, W, H, vMax, vMin, dpr = 1 } = state;
  const ctx = canvas.getContext("2d");
  ctx.setTransform(dpr, 0, 0, dpr, 0, 0);
  ctx.clearRect(0, 0, W, H);

  ctx.strokeStyle = "#f0f2f7"; ctx.lineWidth = 1;
  for (let i = 0; i <= 5; i++) {
    const y = padT + (cH/5)*i, val = vMax - ((vMax - vMin)/5)*i;
    ctx.beginPath(); ctx.moveTo(padL, y); ctx.lineTo(W - padR, y); ctx.stroke();
    ctx.fillStyle = "#aaa"; ctx.font = "10px sans-serif"; ctx.textAlign = "right";
    ctx.fillText("¥" + Math.round(val).toLocaleString(), padL - 4, y + 3);
  }
  ctx.fillStyle = "#aaa"; ctx.font = "10px sans-serif";
  dates.forEach((d, i) => {
    const x = padL + (dates.length > 1 ? i / (dates.length - 1) : 0.5) * cW;
    ctx.save();
    ctx.translate(x, H - padB + 6);
    ctx.rotate(-Math.PI / 4);
    ctx.textAlign = "right";
    ctx.fillText(d.slice(5), 0, 0);
    ctx.restore();
  });

  allPoints.forEach(({ color, pts }) => {
    ctx.strokeStyle = color; ctx.lineWidth = 2; ctx.setLineDash([]);
    ctx.beginPath(); let started = false;
    pts.forEach(p => {
      if (!p) { started = false; return; }
      if (!started) { ctx.moveTo(p.x, p.y); started = true; } else ctx.lineTo(p.x, p.y);
    });
    ctx.stroke();
  });

  allPoints.forEach(({ color, pts }) => {
    pts.forEach(p => {
      if (!p) return;
      if (hoverPt && Math.abs(p.x - hoverPt.x) < 1 && Math.abs(p.y - hoverPt.y) < 1) return;
      ctx.fillStyle = color;
      ctx.beginPath(); ctx.arc(p.x, p.y, 3, 0, Math.PI*2); ctx.fill();
    });
  });

  candidates.forEach(c => {
    ctx.font = "9px sans-serif"; ctx.fillStyle = c.color; ctx.textAlign = "center";
    ctx.fillText(fmtShort(c.v), c.x, c.placedY);
  });

  candidates.forEach(c => {
    const key = c.partner + "|" + dates[c.di];
    if (_comments[key]) {
      ctx.fillStyle = "#f39c12";
      ctx.beginPath(); ctx.arc(c.x + 7, c.y - 7, 5.5, 0, Math.PI * 2); ctx.fill();
      ctx.fillStyle = "#fff"; ctx.font = "bold 7px sans-serif"; ctx.textAlign = "center";
      ctx.fillText("!", c.x + 7, c.y - 4.5);
    }
  });

  if (hoverPt) {
    ctx.fillStyle = hoverPt.color;
    ctx.beginPath(); ctx.arc(hoverPt.x, hoverPt.y, 6, 0, Math.PI*2); ctx.fill();
    ctx.strokeStyle = "#fff"; ctx.lineWidth = 2.5;
    ctx.beginPath(); ctx.arc(hoverPt.x, hoverPt.y, 6, 0, Math.PI*2); ctx.stroke();

    const line1 = hoverPt.partner;
    const line2 = hoverPt.date || "";
    const line3 = "¥" + Math.round(hoverPt.v).toLocaleString();
    const cmtKey = hoverPt.partner + "|" + (hoverPt.date || "");
    const cmt = _comments[cmtKey] || null;
    const TW = 162, TR = 7;
    const TH = cmt ? 90 : 68;
    let tx = hoverPt.x + 14;
    let ty = hoverPt.y - TH - 12;
    if (tx + TW > W - padR + 4) tx = hoverPt.x - TW - 14;
    if (ty < padT - 4) ty = hoverPt.y + 14;

    ctx.save();
    ctx.shadowColor = "rgba(0,0,0,0.18)"; ctx.shadowBlur = 12; ctx.shadowOffsetY = 2;
    ctx.fillStyle = "#fff";
    ctx.beginPath();
    if (ctx.roundRect) { ctx.roundRect(tx, ty, TW, TH, TR); } else { ctx.rect(tx, ty, TW, TH); }
    ctx.fill();
    ctx.restore();

    ctx.fillStyle = hoverPt.color;
    ctx.beginPath();
    if (ctx.roundRect) { ctx.roundRect(tx, ty, 4, TH, [TR, 0, 0, TR]); } else { ctx.rect(tx, ty, 4, TH); }
    ctx.fill();

    ctx.strokeStyle = hoverPt.color; ctx.lineWidth = 1.5;
    ctx.beginPath();
    if (ctx.roundRect) { ctx.roundRect(tx, ty, TW, TH, TR); } else { ctx.rect(tx, ty, TW, TH); }
    ctx.stroke();

    ctx.fillStyle = hoverPt.color; ctx.font = "bold 11px sans-serif"; ctx.textAlign = "left";
    ctx.fillText(line1, tx + 12, ty + 18);
    ctx.fillStyle = "#999"; ctx.font = "10px sans-serif";
    ctx.fillText(line2, tx + 12, ty + 33);
    ctx.fillStyle = "#1a1a2e"; ctx.font = "bold 18px sans-serif";
    ctx.fillText(line3, tx + 12, ty + 57);
    if (cmt) {
      const shortCmt = cmt.length > 24 ? cmt.slice(0, 24) + "…" : cmt;
      ctx.fillStyle = "#f39c12"; ctx.font = "9px sans-serif";
      ctx.fillText("💬 " + shortCmt, tx + 12, ty + 76);
    }
  }
}

init();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
